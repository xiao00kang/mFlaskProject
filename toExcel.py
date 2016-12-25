# encoding=utf-8
import re

import openpyxl
import logging

import shiyebian_worm

URL_SHIYE = 'http://www.shiyebian.net'


def main():
    provinces = shiyebian_worm.get_list_province(shiyebian_worm.download_html(URL_SHIYE))
    wb = openpyxl.Workbook()

    for d in provinces:
        sheet_province = wb.create_sheet(title=d['province'])
        sheet_province['A1'] = '日期'
        sheet_province['B1'] = '招聘信息'
        sheet_province['C1'] = '详细连接'
        url = URL_SHIYE + d['href']
        cent = w_province_or_city(url + 'index.html', sheet_province, 2)
        url += 'index_2.html'
        w_province_or_city(url, sheet_province, cent)
        logging.info(d['province'] + '写入成功')

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    wb.save('事业编招聘汇总.xlsx')


def w_province_or_city(url, sheet, cent):
    try:
        soup = shiyebian_worm.download_html(url)
    except Exception as e:
        logging.info(e)
        return cent
    l = shiyebian_worm.parse_shengshi(soup)
    for i in range(len(l)):
        sheet.cell(row=i + cent, column=1).value = l[i]['date']
        sheet.cell(row=i + cent, column=2).value = l[i]['title']
        sheet.cell(row=i + cent, column=3).value = l[i]['url']
    cent += len(l)
    return cent


if __name__ == '__main__':
    main()
